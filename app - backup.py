import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl.styles import Font, Border, Side, Alignment

st.set_page_config(page_title="Event Reporter Pro", layout="wide")

st.title("📂 Event Data Consolidator (With Preview)")

# --- SIDEBAR ---
uploaded_file = st.sidebar.file_uploader("Upload Portal Data", type=["csv", "xlsx"])
months = ["January", "February", "March", "April", "May", "June", 
          "July", "August", "September", "October", "November", "December"]
now = datetime.now()
selected_month_name = st.sidebar.selectbox("Select Month", months, index=now.month - 1)
selected_month_int = months.index(selected_month_name) + 1
selected_year = st.sidebar.number_input("Year", min_value=2000, max_value=2100, value=now.year)

if uploaded_file:
    # 1. Load Data
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
    df.columns = [c.strip() for c in df.columns]

    date_col = next((c for c in df.columns if c.lower() == "event date"), None)
    status_col = next((c for c in df.columns if c.lower() == "status"), None)
    time_col = next((c for c in df.columns if c.lower() == "time"), None)

    if date_col and status_col:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        filtered_df = df[
            (df[status_col].astype(str).str.strip().str.lower() == "approved") &
            (df[date_col].dt.month == selected_month_int) &
            (df[date_col].dt.year == selected_year)
        ].copy()

        if not filtered_df.empty:
            # 2. FIXED SORTING
            if time_col:
                filtered_df['_tmp_sort'] = pd.to_datetime(filtered_df[time_col].astype(str), errors='coerce').dt.time
                filtered_df = filtered_df.sort_values(by='_tmp_sort').drop(columns=['_tmp_sort'])
            
            filtered_df = filtered_df.reset_index(drop=True)

            # Get Date for Title
            event_date_sample = filtered_df[date_col].iloc[0].strftime('%B %d, %Y')
            title_text = f"Repair Kopitiam@Coral Ris- National Repair Day {event_date_sample}"

            # 3. CONSTRUCTING THE PREVIEW DATAFRAME
            preview_df = pd.DataFrame()
            preview_df['Comment'] = [""] * len(filtered_df)
            preview_df['Q.No'] = ""
            preview_df['S.No'] = range(1, len(filtered_df) + 1)
            preview_df['User'] = filtered_df['User'].values if 'User' in filtered_df.columns else ""
            preview_df['Phone'] = filtered_df['Phone'].values if 'Phone' in filtered_df.columns else ""
            preview_df['Time'] = filtered_df[time_col].values if time_col else ""
            preview_df['Item 1'] = filtered_df['Item 1'].values if 'Item 1' in filtered_df.columns else ""
            preview_df['Item 1 Faults'] = "Not Working"
            preview_df['Item 2'] = filtered_df['Item 2'].values if 'Item 2' in filtered_df.columns else ""
            preview_df['Item 2 Faults'] = "Not Working"
            preview_df['Total Items'] = ""
            preview_df['Items Repaired'] = ""

            # --- PREVIEW SECTION ---
            st.subheader("👀 Data Preview")
            st.info(f"Report Title: **{title_text}**")
            st.dataframe(preview_df, use_container_width=True)
            
            st.write(f"**Total Records Found:** {len(preview_df)}")
            st.write("*(Note: Excel formatting like bold text and borders will be applied during download)*")

            # 4. EXCEL GENERATION (Only happens when download button is clicked)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                preview_df.to_excel(writer, index=False, sheet_name='Report', startrow=1)
                workbook = writer.book
                worksheet = writer.sheets['Report']

                bold_font = Font(bold=True)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                  top=Side(style='thin'), bottom=Side(style='thin'))

                # Title Row
                worksheet.cell(row=1, column=1, value=title_text)
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(preview_df.columns))
                worksheet.cell(row=1, column=1).font = bold_font
                worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='left')

                # Header Row
                for cell in worksheet[2]:
                    cell.font = bold_font

                # Summary Footers
                current_row = worksheet.max_row + 1
                worksheet.cell(row=current_row, column=1, value="Walk-INs")
                current_row += 11 # 10 empty rows + 1
                
                footers = ["Total approved registration for the event", "Walk-in Registrations", 
                           "No show", "Total attended", "Total items Fixed"]
                for label in footers:
                    worksheet.cell(row=current_row, column=1, value=label)
                    current_row += 1

                # Borders
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                             min_col=1, max_col=len(preview_df.columns)):
                    for cell in row:
                        cell.border = thin_border

            # 5. DOWNLOAD BUTTON
            st.download_button(
                label="📥 Download Final Formatted Excel",
                data=output.getvalue(),
                file_name=f"Repair_Day_Report_{selected_month_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("No records found for the selected month.")
else:
    st.info("Upload Portal Data to see a preview of the consolidated report.")