import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import Font, Border, Side, Alignment

# --- 1. EXCEL FORMATTING ENGINE ---
def generate_excel(df, title_text):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Event Report', startrow=1)
        ws = writer.sheets['Event Report']
        
        bold_f = Font(bold=True)
        thin_s = Side(style='thin')
        border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

        # Dynamic Title with Event Date (Row 1)
        ws.cell(row=1, column=1, value=title_text).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        
        # Apply Borders and Bold Header (Row 2)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = border
                if cell.row == 2:
                    cell.font = bold_f

        # --- SUMMARY ROWS ---
        curr_row = ws.max_row + 2
        ws.cell(row=curr_row, column=1, value="Walk-INs").font = bold_f
        
        # Add 10 empty rows for walk-ins with borders
        for r in range(curr_row + 1, curr_row + 11):
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=r, column=c).border = border
        
        # Statistics Footer
        stats_row = curr_row + 11
        labels = [
            "Total approved registration for the event",
            "Walk-in Registrations",
            "No show",
            "Total attended",
            "Total items Fixed"
        ]
        
        for label in labels:
            ws.cell(row=stats_row, column=1, value=label).font = bold_f
            ws.cell(row=stats_row, column=2).border = border 
            stats_row += 1

    return output.getvalue()

# --- 2. STREAMLIT UI ---
st.set_page_config(page_title="RK Coral Ris Event Tool@Bala", layout="wide")
st.title("🛠️ RK Coral Ris Event Tool@Bala")

uploaded_file = st.sidebar.file_uploader("Upload Portal CSV", type=["csv", "xlsx"])
months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
sel_month_name = st.sidebar.selectbox("Select Event Month", months, index=1)
sel_month_int = months.index(sel_month_name) + 1
sel_year = st.sidebar.number_input("Year", value=2026)

if uploaded_file:
    try:
        df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
        df.columns = [str(c).strip() for c in df.columns]

        d_col = 'Event Date'
        s_col = 'status'

        if d_col in df.columns and s_col in df.columns:
            # FORCE DATE FIX: Correcting the Aug/Feb swap found in your file
            df[d_col] = pd.to_datetime(df[d_col], errors='coerce')
            
            def force_date_correction(d):
                if pd.isnull(d): return d
                if d.month == 8 and d.day == 2: # Detected swap
                    return d.replace(month=2, day=8)
                return d
            
            df[d_col] = df[d_col].apply(force_date_correction)

            # Filter for Approved + Selected Month/Year
            mask = (
                (df[s_col].astype(str).str.contains('approve', case=False, na=False)) &
                (df[d_col].dt.month == sel_month_int) &
                (df[d_col].dt.year == sel_year)
            )
            filtered = df[mask].copy()

            if not filtered.empty:
                filtered = filtered.sort_values('Time').reset_index(drop=True)
                
                # Get the event date from the first record for the header
                actual_event_date = filtered[d_col].iloc[0].strftime('%d %B %Y')
                
                # --- 3. CONSTRUCT TABLE ---
                report = pd.DataFrame()
                report['Comment'] = [""] * len(filtered)
                report['Q.No'] = [""] * len(filtered)
                report['S.No'] = range(1, len(filtered) + 1)
                report['User'] = filtered['User']
                report['Phone'] = filtered['Phone']
                report['Time'] = filtered['Time']
                report['Item 1'] = filtered['Item 1']
                report['Item 2'] = filtered['Item 2']
                
                # Sum items per registration
                def get_total(row):
                    count = 0
                    if pd.notnull(row['Item 1']) and str(row['Item 1']).strip() != "": count += 1
                    if pd.notnull(row['Item 2']) and str(row['Item 2']).strip() != "": count += 1
                    return count
                
                report['Total Items'] = report.apply(get_total, axis=1)
                report['Items Repaired'] = [""] * len(filtered)

                # Total Sum for the entire event
                total_event_items = report['Total Items'].sum()

                st.success(f"✅ Records Found: {len(report)} | Total Items to Repair: {total_event_items}")
                st.dataframe(report, use_container_width=True)

                # Use the actual event date in the header
                full_title = f"RK Coral Ris National Repair Day Report - {actual_event_date}"
                
                st.download_button(
                    "📥 Download Final Excel", 
                    generate_excel(report, full_title), 
                    f"RK_Coral Ris Report_{actual_event_date.replace(' ', '_')}.xlsx"
                )
            else:
                st.warning(f"No records found for {sel_month_name} {sel_year}.")
    except Exception as e:
        st.error(f"Error: {e}")